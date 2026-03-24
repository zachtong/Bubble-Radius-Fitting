"""Tests for CSV and Excel export functions."""

import csv
import os

import numpy as np
import pytest

from bubbletrack.model.export import export_csv, export_excel


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def sample_data():
    """Return (radius, centers) with some unprocessed frames."""
    radius = np.array([10.0, -1.0, 20.0, 30.0, -1.0])
    centers = np.array([
        [50.0, 60.0],
        [0.0, 0.0],
        [51.0, 61.0],
        [52.0, 62.0],
        [0.0, 0.0],
    ])
    return radius, centers


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

class TestExportCSV:
    def test_creates_file(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers)
        assert os.path.exists(out)

    def test_skips_unprocessed_frames(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Header + 3 valid frames (indices 0, 2, 3)
        assert len(rows) == 4
        assert rows[0][0] == "Frame"
        # Frame numbers should be 1, 3, 4 (1-indexed)
        assert rows[1][0] == "1"
        assert rows[2][0] == "3"
        assert rows[3][0] == "4"

    def test_basic_columns_no_fps_no_scale(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)

        assert header == ["Frame", "Radius_px", "Center_Row_px", "Center_Col_px"]

    def test_with_fps(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers, fps=1e6)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            first_row = next(reader)

        assert "Time_s" in header
        # Time for frame 0 at 1 MHz should be 0.0
        time_idx = header.index("Time_s")
        assert float(first_row[time_idx]) == pytest.approx(0.0)

    def test_with_scale(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers, um2px=3.2)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            first_row = next(reader)

        assert "Radius_um" in header
        assert "Center_Row_um" in header
        assert "Center_Col_um" in header
        # Radius_um for first valid frame: 10.0 * 3.2 = 32.0
        r_um_idx = header.index("Radius_um")
        assert float(first_row[r_um_idx]) == pytest.approx(32.0)

    def test_with_fps_and_scale(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.csv")
        export_csv(out, radius, centers, fps=1000.0, um2px=2.0)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)

        expected = [
            "Frame", "Time_s",
            "Radius_px", "Center_Row_px", "Center_Col_px",
            "Radius_um", "Center_Row_um", "Center_Col_um",
        ]
        assert header == expected

    def test_all_invalid_produces_header_only(self, tmp_path):
        radius = np.array([-1.0, -1.0, -1.0])
        centers = np.zeros((3, 2))
        out = str(tmp_path / "empty.csv")
        export_csv(out, radius, centers)

        with open(out, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))

        assert len(rows) == 1  # header only

    def test_radius_precision(self, tmp_path):
        radius = np.array([12.34567])
        centers = np.array([[100.98765, 200.12345]])
        out = str(tmp_path / "precision.csv")
        export_csv(out, radius, centers)

        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            row = next(reader)

        # 4 decimal places
        assert row[1] == "12.3457"
        assert row[2] == "100.9877"


# ---------------------------------------------------------------------------
# Excel tests
# ---------------------------------------------------------------------------

class TestExportExcel:
    def test_creates_file(self, tmp_path, sample_data):
        radius, centers = sample_data
        out = str(tmp_path / "test.xlsx")
        export_excel(out, radius, centers)
        assert os.path.exists(out)

    def test_skips_unprocessed_frames(self, tmp_path, sample_data):
        from openpyxl import load_workbook

        radius, centers = sample_data
        out = str(tmp_path / "test.xlsx")
        export_excel(out, radius, centers)

        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Header + 3 valid frames
        assert len(rows) == 4
        assert rows[0][0] == "Frame"
        assert rows[1][0] == "1"
        assert rows[2][0] == "3"
        assert rows[3][0] == "4"

    def test_sheet_title(self, tmp_path, sample_data):
        from openpyxl import load_workbook

        radius, centers = sample_data
        out = str(tmp_path / "test.xlsx")
        export_excel(out, radius, centers)

        wb = load_workbook(out)
        assert wb.active.title == "Bubble Radius Data"

    def test_with_fps_and_scale(self, tmp_path, sample_data):
        from openpyxl import load_workbook

        radius, centers = sample_data
        out = str(tmp_path / "test.xlsx")
        export_excel(out, radius, centers, fps=1000.0, um2px=2.0)

        wb = load_workbook(out)
        ws = wb.active
        header = [cell.value for cell in ws[1]]

        expected = [
            "Frame", "Time_s",
            "Radius_px", "Center_Row_px", "Center_Col_px",
            "Radius_um", "Center_Row_um", "Center_Col_um",
        ]
        assert header == expected

    def test_all_invalid_produces_header_only(self, tmp_path):
        from openpyxl import load_workbook

        radius = np.array([-1.0, -1.0, -1.0])
        centers = np.zeros((3, 2))
        out = str(tmp_path / "empty.xlsx")
        export_excel(out, radius, centers)

        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # header only
